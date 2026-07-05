"""
Administrator Routes Blueprint
Contains customer management, billing management, overdue bill handling,
organization settings, team management, service/parts catalog, and inventory
"""
from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, session, g, current_app, send_from_directory, abort
from datetime import date, timedelta
import logging
from pathlib import Path
import os
from decimal import Decimal, InvalidOperation
from app.services.customer_service import CustomerService
from app.services.job_service import JobService
from app.extensions import db as ext_db
from app.models.job import Job as AdminJob, JobService as JobServiceModel, JobPart
from app.models.tenant import Tenant
from sqlalchemy import and_, or_, exists
from app.services.billing_service import BillingService
from app.utils.decorators import handle_database_errors, log_function_call, validate_pagination
from app.utils.validators import sanitize_input, validate_positive_integer, validate_service_data, validate_part_data
from werkzeug.utils import secure_filename
from werkzeug.datastructures import FileStorage

# Create blueprint
administrator_bp = Blueprint('administrator', __name__)
logger = logging.getLogger(__name__)

# Initialize services
customer_service = CustomerService()
job_service = JobService()
billing_service = BillingService()

DOCUMENT_LIBRARY = [
    {
        'key': '01_Ceg_dokumentumok',
        'label': 'Cég alapdokumentumok',
        'items': [
            'Cégkivonat',
            'Alapító okirat',
            'Adószám igazolás',
            'Bankszámlaszerződés',
            'Kamarai regisztráció',
        ],
    },
    {
        'key': '02_Dolgozok',
        'label': 'Dolgozók',
        'items': [
            'Munkaszerződés',
            'NAV bejelentés (T1041)',
            'Orvosi alkalmassági',
            'Személyi okmányok másolata',
            'Munkaköri leírás',
            'Oktatási papírok',
        ],
    },
    {
        'key': '03_Munkaido',
        'label': 'Munkaidő és jelenlét',
        'items': [
            'Jelenléti ívek (havonta)',
            'Munkaidő nyilvántartás',
            'Szabadság nyilvántartás',
        ],
    },
    {
        'key': '04_Munkavedelem',
        'label': 'Munkavédelem',
        'items': [
            'Munkavédelmi szabályzat',
            'Kockázatértékelés',
            'Védőeszköz szabályzat',
            'Baleseti napló',
        ],
        'subfolders': {
            'Oktatasok': [
                'Oktatási jegyzőkönyvek',
                'Jelenléti ívek',
            ],
        },
    },
    {
        'key': '05_Tuzvedelem',
        'label': 'Tűzvédelem',
        'items': [
            'Tűzvédelmi szabályzat',
            'Tűzoltó készülék ellenőrzések',
            'Oktatási dokumentumok',
        ],
    },
    {
        'key': '06_Szakmai',
        'label': 'Szakmai (villanyszerelés)',
        'items': [
            'Végzettségek',
            'Jogosultságok',
            'Felülvizsgálói papírok',
        ],
        'subfolders': {
            'Jegyzokonyvek': [
                'Érintésvédelem',
                'EPH',
                'Mérési jegyzőkönyvek',
            ],
        },
    },
    {
        'key': '07_Eszkozok',
        'label': 'Eszközök és járművek',
        'items': [
            'Eszköz nyilvántartás',
            'Mérőműszerek hitelesítése',
            'Védőfelszerelések nyilvántartása',
            'Forgalmi',
            'Biztosítás',
            'Cégautó dokumentumok',
        ],
    },
    {
        'key': '08_Konyveles',
        'label': 'Könyvelés / NAV',
        'items': [
            'Számlák (bejövő/kimenő)',
            'ÁFA bevallások',
            'Bérpapírok',
            'Járulék befizetések',
        ],
    },
    {
        'key': '09_Szerzodesek',
        'label': 'Szerződések és munkák',
        'items': [
            'Vállalkozási szerződések',
            'Árajánlatok',
            'Teljesítési igazolások',
        ],
    },
    {
        'key': '10_Epitesi_dok',
        'label': 'Építési dokumentáció',
        'items': [
            'E-napló export',
            'Műszaki dokumentáció',
            'Alvállalkozói szerződések',
        ],
    },
    {
        'key': '11_Biztositasok',
        'label': 'Biztosítások',
        'items': [
            'Felelősségbiztosítás',
            'Kötvények',
            'Befizetések',
        ],
    },
]


def _document_root_for_tenant(tenant: Tenant) -> Path:
    base_root = Path(current_app.instance_path) / 'company_documents'
    tenant_root = base_root / f"tenant_{tenant.tenant_id}_{secure_filename(tenant.slug or tenant.name)}"
    tenant_root.mkdir(parents=True, exist_ok=True)
    return tenant_root


def _ensure_document_library(tenant: Tenant) -> Path:
    tenant_root = _document_root_for_tenant(tenant)
    for section in DOCUMENT_LIBRARY:
        section_root = tenant_root / section['key']
        section_root.mkdir(parents=True, exist_ok=True)
        for subfolder in section.get('subfolders', {}):
            (section_root / subfolder).mkdir(parents=True, exist_ok=True)
    return tenant_root


def _safe_segment(value: str, fallback: str) -> str:
    cleaned = secure_filename((value or '').strip())
    return cleaned or fallback


def _resolve_upload_folder(tenant_root: Path, section_key: str, employee_name: str, project_name: str, accounting_period: str) -> Path:
    target = tenant_root / section_key
    if section_key == '02_Dolgozok':
        target = target / _safe_segment(employee_name, 'Dolgozo')
    elif section_key == '08_Konyveles' and accounting_period:
        period = accounting_period.strip().replace('\\', '/')
        parts = [secure_filename(part) for part in period.split('/') if part.strip()]
        for part in parts:
            if part:
                target = target / part
    elif section_key == '09_Szerzodesek' and project_name:
        target = target / _safe_segment(project_name, 'Projekt')
    target.mkdir(parents=True, exist_ok=True)
    return target


def _build_document_overview(tenant_root: Path) -> list[dict]:
    overview: list[dict] = []
    for section in DOCUMENT_LIBRARY:
        section_root = tenant_root / section['key']
        files = []
        if section_root.exists():
            for file_path in sorted(section_root.rglob('*')):
                if not file_path.is_file():
                    continue
                relative_path = file_path.relative_to(tenant_root).as_posix()
                files.append({
                    'name': file_path.name,
                    'relative_path': relative_path,
                    'folder': file_path.parent.relative_to(tenant_root).as_posix(),
                    'size_kb': round(file_path.stat().st_size / 1024, 1),
                    'updated_at': file_path.stat().st_mtime,
                })
        overview.append({
            'key': section['key'],
            'label': section['label'],
            'required_items': section['items'],
            'subfolders': section.get('subfolders', {}),
            'files': files,
        })
    return overview


def _redirect_back(default_endpoint: str):
    return_to = request.form.get('return_to') or request.args.get('return_to')
    if return_to and return_to.startswith('/'):
        return redirect(return_to)
    referrer = request.referrer or ''
    if referrer.startswith(request.host_url):
        return redirect(referrer)
    return redirect(url_for(default_endpoint))


def _normalize_import_header(value: str) -> str:
    normalized = sanitize_input(value).lower()
    replacements = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ö': 'o', 'ő': 'o',
        'ú': 'u', 'ü': 'u', 'ű': 'u',
    }
    for old, new in replacements.items():
        normalized = normalized.replace(old, new)
    for token in (' ', '-', '/', '\\', '.', '(', ')'):
        normalized = normalized.replace(token, '_')
    while '__' in normalized:
        normalized = normalized.replace('__', '_')
    return normalized.strip('_')


def _parse_money_value(raw_value):
    if raw_value is None:
        return None
    if isinstance(raw_value, (int, float, Decimal)):
        return Decimal(str(raw_value))

    text = sanitize_input(raw_value)
    if not text:
        return None
    text = text.replace('Ft', '').replace('ft', '').replace(' ', '')
    text = text.replace(',', '.')
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _load_import_rows(upload: FileStorage) -> list[dict]:
    filename = (upload.filename or '').lower()
    if filename.endswith('.xlsx'):
        from openpyxl import load_workbook

        workbook = load_workbook(upload, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
    elif filename.endswith('.xls'):
        import xlrd

        workbook = xlrd.open_workbook(file_contents=upload.read())
        sheet = workbook.sheet_by_index(0)
        values = [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)]
    else:
        raise ValueError('Csak .xls vagy .xlsx fájl importálható.')

    if not values:
        return []

    headers = [_normalize_import_header(cell or '') for cell in values[0]]
    rows: list[dict] = []
    for raw_row in values[1:]:
        if not any(cell not in (None, '') for cell in raw_row):
            continue
        row_data = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = raw_row[index] if index < len(raw_row) else None
        rows.append(row_data)
    return rows


def _import_services_from_upload(upload: FileStorage, tenant_id: int) -> dict:
    from app.models.service import Service
    from app.extensions import db

    rows = _load_import_rows(upload)
    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    header_aliases = {
        'service_name': ('service_name', 'name', 'service', 'szolgaltatas', 'szolgaltatas_neve'),
        'cost': ('cost', 'price', 'ar', 'service_cost', 'netto_ar'),
        'category': ('category', 'kategoria'),
        'description': ('description', 'leiras', 'megjegyzes'),
        'estimated_duration_minutes': ('estimated_duration_minutes', 'estimated_duration', 'duration', 'duration_minutes', 'ido', 'ido_perc', 'perc'),
    }

    for row_number, row in enumerate(rows, start=2):
        mapped = {}
        for field_name, aliases in header_aliases.items():
            for alias in aliases:
                if alias in row:
                    mapped[field_name] = row.get(alias)
                    break

        service_name = sanitize_input(mapped.get('service_name', ''))
        cost_value = _parse_money_value(mapped.get('cost'))
        if not service_name:
            skipped += 1
            errors.append(f'{row_number}. sor: hiányzó szolgáltatásnév.')
            continue
        if cost_value is None:
            skipped += 1
            errors.append(f'{row_number}. sor: hibás ár.')
            continue

        existing = db.session.execute(
            db.select(Service).where(
                Service.tenant_id == tenant_id,
                Service.service_name.ilike(service_name)
            )
        ).scalar_one_or_none()

        if existing:
            existing.cost = cost_value
            existing.category = sanitize_input(mapped.get('category', existing.category or '')) or existing.category
            existing.description = sanitize_input(mapped.get('description', existing.description or '')) or existing.description
            duration = mapped.get('estimated_duration_minutes')
            if duration not in (None, ''):
                try:
                    existing.estimated_duration_minutes = int(float(duration))
                except (TypeError, ValueError):
                    pass
            existing.is_active = True
            updated += 1
        else:
            service = Service(
                tenant_id=tenant_id,
                service_name=service_name,
                cost=cost_value,
                category=sanitize_input(mapped.get('category', '')) or 'Általános',
                description=sanitize_input(mapped.get('description', '')),
                is_active=True,
            )
            duration = mapped.get('estimated_duration_minutes')
            if duration not in (None, ''):
                try:
                    service.estimated_duration_minutes = int(float(duration))
                except (TypeError, ValueError):
                    pass
            db.session.add(service)
            created += 1

    db.session.commit()
    return {'created': created, 'updated': updated, 'skipped': skipped, 'errors': errors}


def _import_parts_from_upload(upload: FileStorage, tenant_id: int) -> dict:
    from app.models.part import Part
    from app.extensions import db

    rows = _load_import_rows(upload)
    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    header_aliases = {
        'part_name': ('part_name', 'name', 'part', 'product', 'termek', 'termek_neve'),
        'cost': ('cost', 'price', 'ar', 'part_cost', 'netto_ar'),
        'sku': ('sku', 'cikkszam', 'item_code', 'code'),
        'category': ('category', 'kategoria'),
        'supplier': ('supplier', 'szallito'),
        'description': ('description', 'leiras', 'megjegyzes'),
    }

    for row_number, row in enumerate(rows, start=2):
        mapped = {}
        for field_name, aliases in header_aliases.items():
            for alias in aliases:
                if alias in row:
                    mapped[field_name] = row.get(alias)
                    break

        part_name = sanitize_input(mapped.get('part_name', ''))
        cost_value = _parse_money_value(mapped.get('cost'))
        sku_value = sanitize_input(mapped.get('sku', '')) or None
        if not part_name:
            skipped += 1
            errors.append(f'{row_number}. sor: hiányzó terméknév.')
            continue
        if cost_value is None:
            skipped += 1
            errors.append(f'{row_number}. sor: hibás ár.')
            continue

        existing = None
        if sku_value:
            existing = db.session.execute(
                db.select(Part).where(
                    Part.tenant_id == tenant_id,
                    Part.sku == sku_value
                )
            ).scalar_one_or_none()
        if not existing:
            existing = db.session.execute(
                db.select(Part).where(
                    Part.tenant_id == tenant_id,
                    Part.part_name.ilike(part_name)
                )
            ).scalar_one_or_none()

        if existing:
            existing.part_name = part_name
            existing.cost = cost_value
            existing.sku = sku_value
            existing.category = sanitize_input(mapped.get('category', existing.category or '')) or existing.category
            existing.supplier = sanitize_input(mapped.get('supplier', existing.supplier or '')) or existing.supplier
            existing.description = sanitize_input(mapped.get('description', existing.description or '')) or existing.description
            existing.is_active = True
            updated += 1
        else:
            part = Part(
                tenant_id=tenant_id,
                part_name=part_name,
                cost=cost_value,
                sku=sku_value,
                category=sanitize_input(mapped.get('category', '')) or 'Általános',
                supplier=sanitize_input(mapped.get('supplier', '')) or None,
                description=sanitize_input(mapped.get('description', '')),
                is_active=True,
            )
            db.session.add(part)
            created += 1

    db.session.commit()
    return {'created': created, 'updated': updated, 'skipped': skipped, 'errors': errors}


def require_admin_login():
    """Check administrator login status"""
    if not session.get('logged_in'):
        flash('Please login first', 'warning')
        return redirect(url_for('auth.login'))

    if session.get('current_role') not in ('owner', 'admin'):
        flash('Administrator privileges required', 'error')
        return redirect(url_for('main.index'))

    return None


@administrator_bp.route('/dashboard')
@handle_database_errors
@log_function_call
def dashboard():
    """Administrator dashboard"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    try:
        # Get system statistics
        job_stats = job_service.get_job_statistics()
        billing_stats = billing_service.get_billing_statistics()

        # Get customer statistics
        total_customers = len(customer_service.get_all_customers())
        customers_with_unpaid = customer_service.get_customers_with_filter(has_unpaid=True)
        customers_with_overdue = customer_service.get_customers_with_filter(has_overdue=True)

        # Get recent activities
        recent_jobs, _, _ = job_service.get_current_jobs(page=1, per_page=5)

        from app.models.job import Job as _Job
        from flask import session as _session
        _tenant_id = _session.get('current_tenant_id') or 1

        # Havi bevétel az aktuális évre
        _Job2 = _Job
        current_year = date.today().year
        monthly_revenue = [0.0] * 12
        monthly_q = ext_db.select(
            ext_db.func.extract('month', _Job2.job_date).label('month'),
            ext_db.func.coalesce(ext_db.func.sum(_Job2.total_cost), 0).label('revenue')
        ).where(
            and_(
                ext_db.func.extract('year', _Job2.job_date) == current_year,
                _Job2.tenant_id == _tenant_id,
                _Job2.total_cost > 0
            )
        ).group_by(ext_db.func.extract('month', _Job2.job_date))
        for row in ext_db.session.execute(monthly_q):
            monthly_revenue[int(row.month) - 1] = float(row.revenue)

        billing_stats['monthly_revenue'] = monthly_revenue

        _overdue_q = ext_db.select(_Job).where(
            and_(_Job.completed == True, _Job.paid == False, _Job.tenant_id == _tenant_id)
        ).order_by(_Job.job_id.desc()).limit(5)
        overdue_bills = list(ext_db.session.execute(_overdue_q).scalars())

        return render_template('administrator/dashboard.html',
                             job_stats=job_stats,
                             billing_stats=billing_stats,
                             total_customers=total_customers,
                             customers_with_unpaid=len(customers_with_unpaid),
                             customers_with_overdue=len(customers_with_overdue),
                             recent_jobs=recent_jobs,
                             overdue_bills=overdue_bills,
                             current_date=date.today())

    except Exception as e:
        logger.error(f"Administrator dashboard loading failed: {e}")
        flash('Failed to load dashboard', 'error')
        return render_template('administrator/dashboard.html',
                             job_stats={},
                             billing_stats={},
                             total_customers=0,
                             customers_with_unpaid=0,
                             customers_with_overdue=0,
                             recent_jobs=[],
                             overdue_bills=[],
                             current_date=date.today())


@administrator_bp.route('/documents', methods=['GET', 'POST'])
@log_function_call
def document_center():
    """Company document center with predefined compliance folders."""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    tenant_id = session.get('current_tenant_id') or getattr(g, 'current_tenant_id', None)
    tenant = ext_db.session.get(Tenant, tenant_id) if tenant_id else None
    if not tenant:
        flash('Szervezet nem található a dokumentumtárhoz.', 'error')
        return redirect(url_for('administrator.dashboard'))

    tenant_root = _ensure_document_library(tenant)

    if request.method == 'POST':
        action = request.form.get('action', 'upload')

        if action == 'create_folder':
            section_key = sanitize_input(request.form.get('section_key', ''))
            employee_name = sanitize_input(request.form.get('employee_name', ''))
            project_name = sanitize_input(request.form.get('project_name', ''))
            accounting_year = sanitize_input(request.form.get('accounting_year', ''))
            accounting_month = sanitize_input(request.form.get('accounting_month', ''))

            if section_key not in {section['key'] for section in DOCUMENT_LIBRARY}:
                flash('Ismeretlen dokumentum szekció.', 'error')
                return redirect(url_for('administrator.document_center'))

            accounting_period = ''
            if section_key == '08_Konyveles' and accounting_year and accounting_month:
                accounting_period = f'{accounting_year}/{accounting_month.zfill(2)}'

            _resolve_upload_folder(
                tenant_root,
                section_key,
                employee_name=employee_name,
                project_name=project_name,
                accounting_period=accounting_period,
            )
            flash('A kért almappa létrejött.', 'success')
            return redirect(url_for('administrator.document_center'))

        section_key = sanitize_input(request.form.get('section_key', ''))
        employee_name = sanitize_input(request.form.get('employee_name', ''))
        project_name = sanitize_input(request.form.get('project_name', ''))
        accounting_period = sanitize_input(request.form.get('accounting_period', ''))
        document_label = sanitize_input(request.form.get('document_label', ''))
        upload = request.files.get('file')

        if section_key not in {section['key'] for section in DOCUMENT_LIBRARY}:
            flash('Ismeretlen dokumentum szekció.', 'error')
            return redirect(url_for('administrator.document_center'))
        if not upload or not upload.filename:
            flash('Válassz ki egy feltöltendő fájlt.', 'error')
            return redirect(url_for('administrator.document_center'))

        target_folder = _resolve_upload_folder(
            tenant_root,
            section_key,
            employee_name=employee_name,
            project_name=project_name,
            accounting_period=accounting_period,
        )

        original_name = secure_filename(upload.filename)
        prefix = _safe_segment(document_label, '')
        filename = f'{prefix}_{original_name}' if prefix else original_name
        destination = target_folder / filename
        upload.save(destination)
        flash('Dokumentum sikeresen feltöltve.', 'success')
        return redirect(url_for('administrator.document_center'))

    document_sections = _build_document_overview(tenant_root)
    return render_template(
        'administrator/documents.html',
        document_sections=document_sections,
        tenant=tenant,
    )


@administrator_bp.route('/documents/download/<path:relative_path>')
@log_function_call
def download_document(relative_path: str):
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    tenant_id = session.get('current_tenant_id') or getattr(g, 'current_tenant_id', None)
    tenant = ext_db.session.get(Tenant, tenant_id) if tenant_id else None
    if not tenant:
        abort(404)

    tenant_root = _ensure_document_library(tenant)
    requested_path = (tenant_root / relative_path).resolve()
    if tenant_root.resolve() not in requested_path.parents and requested_path != tenant_root.resolve():
        abort(403)
    if not requested_path.exists() or not requested_path.is_file():
        abort(404)

    return send_from_directory(requested_path.parent, requested_path.name, as_attachment=False)


@administrator_bp.route('/customers')
@validate_pagination
@log_function_call
def customer_list(page=1, per_page=20):
    """Customer management page"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    try:
        # Get filter parameters
        filter_type = sanitize_input(request.args.get('filter', 'all'))
        search_query = sanitize_input(request.args.get('search', ''))

        # Get customers based on filter
        if filter_type == 'unpaid':
            customers = customer_service.get_customers_with_filter(has_unpaid=True)
        elif filter_type == 'overdue':
            customers = customer_service.get_customers_with_filter(has_overdue=True)
        elif search_query:
            customers_obj = customer_service.search_customers(search_query)
            customers = [c.to_dict() for c in customers_obj]
            # Add statistics info
            for customer in customers:
                customer['total_unpaid'] = customer_service.get_customer_by_id(customer['customer_id']).get_total_unpaid_amount()
                customer['has_overdue'] = customer_service.get_customer_by_id(customer['customer_id']).has_overdue_bills()
        else:
            customers_obj = customer_service.get_all_customers()
            customers = []
            for c in customers_obj:
                customer_data = c.to_dict()
                customer_data['total_unpaid'] = c.get_total_unpaid_amount()
                customer_data['has_overdue'] = c.has_overdue_bills()
                customers.append(customer_data)

        # Simple pagination
        total = len(customers)
        start = (page - 1) * per_page
        end = start + per_page
        customers_page = customers[start:end]
        total_pages = (total + per_page - 1) // per_page

        return render_template('administrator/customer_list.html',
                             customers=customers_page,
                             page=page,
                             per_page=per_page,
                             total=total,
                             total_pages=total_pages,
                             filter_type=filter_type,
                             search_query=search_query)

    except Exception as e:
        logger.error(f"Customer management page loading failed: {e}")
        flash('Failed to load customer list', 'error')
        return render_template('administrator/customer_list.html',
                             customers=[],
                             page=1,
                             per_page=per_page,
                             total=0,
                             total_pages=0,
                             filter_type='all',
                             search_query='')


@administrator_bp.route('/billing')
@handle_database_errors
@log_function_call
def billing_management():
    """Billing management page"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    try:
        # Get filter parameters
        filter_type = sanitize_input(request.args.get('filter', 'unpaid'))
        customer_name = sanitize_input(request.args.get('customer', ''))

        # Get billing data
        if filter_type == 'overdue':
            bills = billing_service.get_overdue_bills()
        elif filter_type == 'all':
            bills = billing_service.get_all_bills_with_status()
        else:  # unpaid
            bills = billing_service.get_unpaid_bills(customer_name if customer_name != 'Choose...' else None)

        # Get customer name list for dropdown
        customers = customer_service.get_all_customers()
        customer_names = [f"{c.first_name} {c.family_name}".strip() for c in customers]
        customer_names = list(set(customer_names))  # Remove duplicates
        customer_names.sort()

        # Get billing statistics
        billing_stats = billing_service.get_billing_statistics()

        return render_template('administrator/billing.html',
                             bills=bills,
                             filter_type=filter_type,
                             customer_name=customer_name,
                             customer_names=customer_names,
                             billing_stats=billing_stats)

    except Exception as e:
        logger.error(f"Billing management page loading failed: {e}")
        flash('Failed to load billing management page', 'error')
        return render_template('administrator/billing.html',
                             bills=[],
                             filter_type='unpaid',
                             customer_name='',
                             customer_names=[],
                             billing_stats={})


@administrator_bp.route('/overdue-bills')
@handle_database_errors
@log_function_call
def overdue_bills():
    """Overdue bills - completed jobs with services and/or parts"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    try:
        from app.extensions import db
        from app.models.job import Job, JobService, JobPart
        from flask import session

        tenant_id = session.get('current_tenant_id') or 1

        # Befejezett munkák amiknek van legalább 1 szolgáltatása VAGY Terméke
        has_service = exists().where(JobService.job_id == Job.job_id)
        has_part = exists().where(JobPart.job_id == Job.job_id)

        query = db.select(Job).where(
            and_(
                Job.completed == True,
                Job.tenant_id == tenant_id,
                or_(has_service, has_part)
            )
        ).order_by(Job.job_id.desc())

        overdue_bills_list = list(db.session.execute(query).scalars())
        total_overdue_amount = sum(float(j.total_cost or 0) for j in overdue_bills_list)

        return render_template('administrator/overdue_bills.html',
                             overdue_bills=overdue_bills_list,
                             total_overdue_amount=total_overdue_amount,
                             days_threshold=14,
                             total_count=len(overdue_bills_list))

    except Exception as e:
        logger.error(f"Overdue bills page loading failed: {e}")
        flash(f'Hiba: {str(e)}', 'error')
        return render_template('administrator/overdue_bills.html',
                             overdue_bills=[],
                             total_overdue_amount=0,
                             days_threshold=14,
                             total_count=0)


@administrator_bp.route('/pay-bills')
@handle_database_errors
@log_function_call
def pay_bills():
    """Payment processing page - shows completed jobs"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    try:
        customer_name = sanitize_input(request.args.get('customer', ''))
        completed_jobs = billing_service.get_completed_jobs()

        # Filter by customer name if provided
        if customer_name and customer_name != 'Choose...':
            completed_jobs = [
                j for j in completed_jobs
                if j.customer_rel and
                f"{j.customer_rel.first_name} {j.customer_rel.family_name}".strip() == customer_name
            ]

        customers = customer_service.get_all_customers()
        customer_names = sorted(set(
            f"{c.first_name} {c.family_name}".strip() for c in customers
        ))

        return render_template('administrator/pay_bills.html',
                             unpaid_bills=completed_jobs,
                             customer_name=customer_name,
                             customer_names=customer_names)

    except Exception as e:
        logger.error(f"Payment processing page loading failed: {e}")
        flash('Failed to load payment processing page', 'error')
        return render_template('administrator/pay_bills.html',
                             unpaid_bills=[],
                             customer_name='',
                             customer_names=[])


@administrator_bp.route('/customers/<int:customer_id>/pay', methods=['POST'])
@handle_database_errors
def pay_customer_bills(customer_id):
    """Mark all bills for a customer as paid"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    try:
        success, errors, count = billing_service.mark_customer_bills_as_paid(customer_id)

        if success:
            flash(f'Successfully marked {count} bills as paid!', 'success')
        else:
            for error in errors:
                flash(error, 'error')

        return redirect(url_for('administrator.customer_list'))

    except Exception as e:
        logger.error(f"Failed to mark customer bills as paid: {e}")
        flash('Failed to mark payment, please try again later', 'error')
        return redirect(url_for('administrator.customer_list'))


@administrator_bp.route('/jobs/<int:job_id>/pay', methods=['POST'])
@handle_database_errors
def pay_single_bill(job_id):
    """Mark single work order as paid"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    try:
        success, errors = billing_service.mark_job_as_paid(job_id)

        if success:
            flash('Bill has been marked as paid!', 'success')
        else:
            for error in errors:
                flash(error, 'error')

        # Redirect based on source page
        return_page = sanitize_input(request.form.get('return_page', 'pay_bills'))
        if return_page == 'overdue_bills':
            return redirect(url_for('administrator.overdue_bills'))
        else:
            return redirect(url_for('administrator.pay_bills'))

    except Exception as e:
        logger.error(f"Failed to mark bill as paid: {e}")
        flash('Failed to mark payment, please try again later', 'error')
        return redirect(url_for('administrator.pay_bills'))


@administrator_bp.route('/reports')
@handle_database_errors
@log_function_call
def reports():
    """Reports page"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    try:
        # Get various statistics
        job_stats = job_service.get_job_statistics()
        billing_stats = billing_service.get_billing_statistics()

        # Get customer statistics
        total_customers = len(customer_service.get_all_customers())
        customers_with_unpaid = customer_service.get_customers_with_filter(has_unpaid=True)
        customers_with_overdue = customer_service.get_customers_with_filter(has_overdue=True)

        # Calculate current month vs last month comparison data
        today = date.today()
        current_month_start = today.replace(day=1)
        last_month_end = current_month_start - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)

        report_data = {
            'job_stats': job_stats,
            'billing_stats': billing_stats,
            'customer_stats': {
                'total_customers': total_customers,
                'customers_with_unpaid': len(customers_with_unpaid),
                'customers_with_overdue': len(customers_with_overdue),
                'customer_payment_rate': ((total_customers - len(customers_with_unpaid)) / total_customers * 100) if total_customers > 0 else 0
            },
            'period_info': {
                'current_month': current_month_start.strftime('%B %Y'),
                'last_month': last_month_start.strftime('%B %Y'),
                'generated_date': today.strftime('%Y-%m-%d')
            }
        }

        # Havi bevétel + kiadás az utolsó 6 hónapra
        import calendar
        from app.models.job import Job as JobModel
        from app.extensions import db
        from sqlalchemy import func, extract

        tenant_id = session.get('current_tenant_id') or 1
        monthly_revenue_data = []
        customer_activity_data = []

        for i in range(5, -1, -1):
            ref = today.replace(day=1) - timedelta(days=1)
            # i hónappal ezelőtti hónap első napja
            month_date = (today.replace(day=1) - timedelta(days=1))
            # Számítsuk ki az i-edik hónapot visszafelé
            year = today.year
            month = today.month - i
            while month <= 0:
                month += 12
                year -= 1

            rev = db.session.execute(
                db.select(func.coalesce(func.sum(JobModel.total_cost), 0))
                .where(and_(
                    JobModel.tenant_id == tenant_id,
                    JobModel.completed == True,
                    extract('year', JobModel.job_date) == year,
                    extract('month', JobModel.job_date) == month,
                ))
            ).scalar() or 0

            job_count = db.session.execute(
                db.select(func.count())
                .select_from(JobModel)
                .where(and_(
                    JobModel.tenant_id == tenant_id,
                    extract('year', JobModel.job_date) == year,
                    extract('month', JobModel.job_date) == month,
                ))
            ).scalar() or 0

            try:
                from app.models.expense import Expense
                exp = db.session.execute(
                    db.select(func.coalesce(func.sum(Expense.amount), 0))
                    .where(and_(
                        Expense.tenant_id == tenant_id,
                        extract('year', Expense.expense_date) == year,
                        extract('month', Expense.expense_date) == month,
                    ))
                ).scalar() or 0
            except Exception:
                exp = 0

            try:
                from app.models.customer import Customer
                new_customers = db.session.execute(
                    db.select(func.count())
                    .select_from(Customer)
                    .where(and_(
                        Customer.tenant_id == tenant_id,
                        extract('year', Customer.created_at) == year,
                        extract('month', Customer.created_at) == month,
                    ))
                ).scalar() or 0
            except Exception:
                new_customers = 0

            label = date(year, month, 1).strftime('%b')
            monthly_revenue_data.append({'label': label, 'revenue': float(rev), 'expense': float(exp)})
            customer_activity_data.append({'label': label, 'new_customers': new_customers, 'jobs': job_count})

        return render_template('administrator/reports.html',
                             report_data=report_data,
                             monthly_revenue_data=monthly_revenue_data,
                             customer_activity_data=customer_activity_data)

    except Exception as e:
        logger.error(f"Reports page loading failed: {e}")
        flash('Failed to load reports', 'error')
        return render_template('administrator/reports.html',
                             report_data={})


# API endpoints
@administrator_bp.route('/api/customers/<int:customer_id>/billing-summary')
@handle_database_errors
def api_customer_billing_summary(customer_id):
    """API: Get customer billing summary"""
    try:
        summary = billing_service.get_customer_billing_summary(customer_id)
        return jsonify(summary)

    except Exception as e:
        logger.error(f"Failed to get customer billing summary: {e}")
        return jsonify({'error': 'Failed to get billing summary'}), 500


@administrator_bp.route('/api/billing/statistics')
@handle_database_errors
def api_billing_statistics():
    """API: Get billing statistics"""
    try:
        stats = billing_service.get_billing_statistics()
        return jsonify(stats)

    except Exception as e:
        logger.error(f"Failed to get billing statistics: {e}")
        return jsonify({'error': 'Failed to get statistics'}), 500


@administrator_bp.route('/api/dashboard/summary')
@handle_database_errors
def api_dashboard_summary():
    """API: Get dashboard summary"""
    try:
        job_stats = job_service.get_job_statistics()
        billing_stats = billing_service.get_billing_statistics()

        # Customer statistics
        total_customers = len(customer_service.get_all_customers())
        customers_with_unpaid = customer_service.get_customers_with_filter(has_unpaid=True)
        customers_with_overdue = customer_service.get_customers_with_filter(has_overdue=True)

        summary = {
            'jobs': job_stats,
            'billing': billing_stats,
            'customers': {
                'total': total_customers,
                'with_unpaid': len(customers_with_unpaid),
                'with_overdue': len(customers_with_overdue)
            },
            'alerts': {
                'overdue_bills': len(billing_service.get_overdue_bills()),
                'pending_jobs': job_stats.get('pending_jobs', 0)
            }
        }

        return jsonify(summary)

    except Exception as e:
        logger.error(f"Failed to get dashboard summary: {e}")
        return jsonify({'error': 'Failed to get summary'}), 500


@administrator_bp.route('/api/export/customers')
@handle_database_errors
def api_export_customers():
    """API: Export customer data"""
    try:
        customers = customer_service.get_all_customers()
        customer_data = []

        for c in customers:
            customer_info = c.to_dict()
            customer_info['total_unpaid'] = c.get_total_unpaid_amount()
            customer_info['has_overdue'] = c.has_overdue_bills()
            customer_data.append(customer_info)

        return jsonify({
            'data': customer_data,
            'export_date': date.today().isoformat(),
            'total_count': len(customer_data)
        })

    except Exception as e:
        logger.error(f"Failed to export customer data: {e}")
        return jsonify({'error': 'Failed to export data'}), 500


@administrator_bp.route('/api/customers/<int:customer_id>/summary')
@handle_database_errors
def api_customer_summary(customer_id):
    """API: Get customer summary"""
    try:
        customer = customer_service.get_customer_by_id(customer_id)
        if not customer:
            return jsonify({'error': 'Customer not found'}), 404

        stats = customer_service.get_customer_statistics(customer_id)
        return jsonify(stats)

    except Exception as e:
        logger.error(f"Failed to get customer summary: {e}")
        return jsonify({'error': 'Failed to get customer information'}), 500


# =============================================================================
# ORGANIZATION SETTINGS
# =============================================================================

@administrator_bp.route('/settings', methods=['GET', 'POST'])
@handle_database_errors
@log_function_call
def org_settings():
    """Organization settings page"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    from app.models.tenant import Tenant
    from app.extensions import db

    tenant_id = session.get('current_tenant_id') or getattr(g, 'current_tenant_id', None)
    if not tenant_id:
        flash('No organization selected', 'error')
        return redirect(url_for('main.dashboard'))

    tenant = Tenant.find_by_id(tenant_id)
    if not tenant:
        flash('Organization not found', 'error')
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        try:
            # Handle logo upload as base64
            import base64
            logo_url = tenant.logo_url
            if 'logo_file' in request.files:
                logo_file = request.files['logo_file']
                if logo_file and logo_file.filename:
                    ext = logo_file.filename.rsplit('.', 1)[-1].lower()
                    if ext in ['png', 'jpg', 'jpeg', 'gif', 'svg']:
                        img_data = base64.b64encode(logo_file.read()).decode('utf-8')
                        mime = 'image/svg+xml' if ext == 'svg' else f'image/{ext}'
                        logo_url = f'data:{mime};base64,{img_data}'
                    if ext in ['png', 'jpg', 'jpeg', 'gif', 'svg']:
                        filename = f'logo_{tenant_id}_{uuid.uuid4().hex[:8]}.{ext}'
                        upload_path = os.path.join('app', 'static', 'images', filename)
                        logo_file.save(upload_path)
                        logo_url = f'/static/images/{filename}'

            org_name = request.form.get('org_name', '').strip() or request.form.get('name', '').strip()
            if org_name:
                tenant.name = org_name
            tenant.email = request.form.get('email', '').strip() or tenant.email
            tenant.phone = request.form.get('phone', '').strip() or tenant.phone
            tenant.address = request.form.get('address', '').strip() or tenant.address
            tenant.logo_url = logo_url

            from sqlalchemy.orm.attributes import flag_modified
            settings = dict(tenant.settings or {})
            tax_rate = request.form.get('tax_rate')
            if tax_rate:
                try:
                    settings['tax_rate'] = float(tax_rate)
                except ValueError:
                    pass
            settings['currency'] = request.form.get('currency', 'HUF')
            for field in ['tax_id', 'eu_tax_id', 'bank_account', 'company_reg']:
                val = request.form.get(field, '').strip()
                if val:
                    settings[field] = val
            tenant.settings = settings
            flag_modified(tenant, 'settings')

            session['current_tenant_name'] = tenant.name
            db.session.commit()
            flash('Beállítások mentve!', 'success')
        except Exception as e:
            logger.error(f"Failed to update org settings: {e}")
            db.session.rollback()
            flash('Hiba a mentés során', 'error')

    return render_template('administrator/org_settings.html', tenant=tenant, org=tenant)


# =============================================================================
# TEAM MANAGEMENT
# =============================================================================

@administrator_bp.route('/team')
@handle_database_errors
@log_function_call
def team_members():
    """Team member management page"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    from app.models.tenant_membership import TenantMembership
    from app.models.user import User
    from app.extensions import db

    tenant_id = session.get('current_tenant_id') or getattr(g, 'current_tenant_id', None)
    if not tenant_id:
        flash('No organization selected', 'error')
        return redirect(url_for('main.dashboard'))

    try:
        memberships = db.session.execute(
            db.select(TenantMembership).where(
                TenantMembership.tenant_id == tenant_id
            ).order_by(TenantMembership.role, TenantMembership.created_at)
        ).scalars().all()

        members = []
        for m in memberships:
            user = User.find_by_id(m.user_id)
            if user:
                members.append({
                    'membership_id': m.id,
                    'user_id': m.user_id,
                    'username': user.username,
                    'email': user.email,
                    'role': m.role,
                    'status': m.status,
                    'is_default': m.is_default,
                    'accepted_at': m.accepted_at,
                    'invited_at': m.invited_at,
                })

        return render_template('administrator/team_members.html',
                             members=members,
                             available_roles=TenantMembership.VALID_ROLES)

    except Exception as e:
        logger.error(f"Failed to load team members: {e}")
        flash('Failed to load team members', 'error')
        return render_template('administrator/team_members.html',
                             members=[],
                             available_roles=[])


@administrator_bp.route('/team/invite', methods=['POST'])
@handle_database_errors
def invite_team_member():
    """Invite a new team member"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    from app.services.tenant_service import TenantService

    tenant_id = session.get('current_tenant_id') or getattr(g, 'current_tenant_id', None)
    if not tenant_id:
        flash('No organization selected', 'error')
        return redirect(url_for('main.dashboard'))

    email = sanitize_input(request.form.get('email', ''))
    role = sanitize_input(request.form.get('role', 'viewer'))
    user_id = session.get('user_id')

    if not email:
        flash('Email is required', 'error')
        return redirect(url_for('administrator.team_members'))

    tenant_service = TenantService()
    success, errors, membership = tenant_service.invite_member(
        tenant_id=tenant_id,
        email=email,
        role=role,
        invited_by_user_id=user_id,
    )

    if success:
        flash(f'Invitation sent to {email}!', 'success')
    else:
        for error in errors:
            flash(error, 'error')

    return redirect(url_for('administrator.team_members'))


# =============================================================================
# SERVICE CATALOG
# =============================================================================

@administrator_bp.route('/services', methods=['GET', 'POST'])
@handle_database_errors
@log_function_call
def service_catalog():
    """Service catalog management"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    from app.models.service import Service
    from app.extensions import db

    tenant_id = session.get('current_tenant_id') or getattr(g, 'current_tenant_id', None)

    if request.method == 'POST':
        action = request.form.get('action', 'add')

        if action == 'add':
            data = {
                'service_name': sanitize_input(request.form.get('service_name', '')),
                'cost': request.form.get('cost'),
            }
            validation = validate_service_data(data)
            if not validation.is_valid:
                for error in validation.get_errors():
                    flash(error, 'error')
            else:
                try:
                    service = Service(
                        tenant_id=tenant_id,
                        service_name=data['service_name'],
                        cost=float(data['cost']),
                        category=sanitize_input(request.form.get('category', 'General')),
                        description=sanitize_input(request.form.get('description', '')),
                        estimated_duration_minutes=request.form.get('estimated_duration', type=int),
                        is_active=True,
                    )
                    db.session.add(service)
                    db.session.commit()
                    flash(f'Service "{data["service_name"]}" added!', 'success')
                except Exception as e:
                    logger.error(f"Failed to add service: {e}")
                    db.session.rollback()
                    flash('Failed to add service', 'error')

        elif action == 'import':
            upload = request.files.get('import_file')
            if not upload or not upload.filename:
                flash('Válassz ki egy import fájlt.', 'error')
            else:
                try:
                    result = _import_services_from_upload(upload, tenant_id)
                    flash(
                        f'Szolgáltatás import kész: {result["created"]} új, {result["updated"]} frissített, {result["skipped"]} kihagyott sor.',
                        'success'
                    )
                    for error in result['errors'][:8]:
                        flash(error, 'warning')
                except Exception as e:
                    logger.error(f"Failed to import services: {e}")
                    db.session.rollback()
                    flash('A szolgáltatás import sikertelen.', 'error')

        elif action == 'edit':
            service_id = request.form.get('service_id', type=int)
            if service_id:
                service = Service.find_by_id(service_id)
                if service:
                    service.service_name = sanitize_input(request.form.get('service_name', service.service_name))
                    service.category = sanitize_input(request.form.get('category', service.category))
                    try:
                        service.cost = float(request.form.get('cost', service.cost))
                    except (ValueError, TypeError):
                        pass
                    dur = request.form.get('estimated_duration')
                    if dur:
                        try:
                            service.estimated_duration_minutes = int(dur)
                        except (ValueError, TypeError):
                            pass
                    db.session.commit()
                    flash('Szolgáltatás frissítve!', 'success')

        elif action == 'toggle':
            service_id = request.form.get('service_id', type=int)
            if service_id:
                service = Service.find_by_id(service_id)
                if service:
                    service.is_active = not service.is_active
                    db.session.commit()
                    status = 'activated' if service.is_active else 'deactivated'
                    flash(f'Service {status}!', 'success')

        return _redirect_back('administrator.service_catalog')

    # GET - load services
    try:
        from flask import g
        g.current_tenant_id = tenant_id
        services = Service.get_all_sorted()
        return render_template('administrator/service_catalog.html', services=services)
    except Exception as e:
        logger.error(f"Failed to load service catalog: {e}")
        flash('Failed to load service catalog', 'error')
        return render_template('administrator/service_catalog.html', services=[])


# =============================================================================
# PARTS CATALOG
# =============================================================================

@administrator_bp.route('/parts', methods=['GET', 'POST'])
@handle_database_errors
@log_function_call
def parts_catalog():
    """Parts catalog management"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    from app.models.part import Part
    from app.extensions import db

    tenant_id = session.get('current_tenant_id') or getattr(g, 'current_tenant_id', None)

    if request.method == 'POST':
        action = request.form.get('action', 'add')

        if action == 'add':
            data = {
                'part_name': sanitize_input(request.form.get('part_name', '')),
                'cost': request.form.get('cost'),
            }
            validation = validate_part_data(data)
            if not validation.is_valid:
                for error in validation.get_errors():
                    flash(error, 'error')
            else:
                try:
                    part = Part(
                        tenant_id=tenant_id,
                        part_name=data['part_name'],
                        cost=float(data['cost']),
                        sku=sanitize_input(request.form.get('sku', '')) or None,
                        category=sanitize_input(request.form.get('category', 'General')),
                        description=sanitize_input(request.form.get('description', '')),
                        supplier=sanitize_input(request.form.get('supplier', '')) or None,
                        is_active=True,
                    )
                    db.session.add(part)
                    db.session.commit()
                    flash(f'Part "{data["part_name"]}" added!', 'success')
                except Exception as e:
                    logger.error(f"Failed to add part: {e}")
                    db.session.rollback()
                    flash('Failed to add part', 'error')

        elif action == 'import':
            upload = request.files.get('import_file')
            if not upload or not upload.filename:
                flash('Válassz ki egy import fájlt.', 'error')
            else:
                try:
                    result = _import_parts_from_upload(upload, tenant_id)
                    flash(
                        f'Termék import kész: {result["created"]} új, {result["updated"]} frissített, {result["skipped"]} kihagyott sor.',
                        'success'
                    )
                    for error in result['errors'][:8]:
                        flash(error, 'warning')
                except Exception as e:
                    logger.error(f"Failed to import parts: {e}")
                    db.session.rollback()
                    flash('A termék import sikertelen.', 'error')

        elif action == 'edit':
            part_id = request.form.get('part_id', type=int)
            if part_id:
                part = Part.find_by_id(part_id)
                if part:
                    part.part_name = sanitize_input(request.form.get('part_name', part.part_name))
                    part.category = sanitize_input(request.form.get('category', part.category))
                    part.supplier = sanitize_input(request.form.get('supplier', ''))
                    try:
                        part.cost = float(request.form.get('cost', part.cost))
                    except (ValueError, TypeError):
                        pass
                    part.sku = sanitize_input(request.form.get('sku', ''))
                    db.session.commit()
                    flash('Termék frissítve!', 'success')

        elif action == 'toggle':
            part_id = request.form.get('part_id', type=int)
            if part_id:
                part = Part.find_by_id(part_id)
                if part:
                    part.is_active = not part.is_active
                    db.session.commit()
                    status = 'activated' if part.is_active else 'deactivated'
                    flash(f'Part {status}!', 'success')

        return _redirect_back('administrator.parts_catalog')

    # GET - load parts
    try:
        from flask import g
        g.current_tenant_id = tenant_id
        parts = Part.get_all_sorted()
        return render_template('administrator/parts_catalog.html', parts=parts)
    except Exception as e:
        logger.error(f"Failed to load parts catalog: {e}")
        flash('Failed to load parts catalog', 'error')
        return render_template('administrator/parts_catalog.html', parts=[])


# =============================================================================
# INVENTORY MANAGEMENT
# =============================================================================

@administrator_bp.route('/inventory')
@handle_database_errors
@log_function_call
def inventory():
    """Inventory dashboard"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    from app.models.inventory import Inventory, InventoryTransaction
    from app.extensions import db

    tenant_id = session.get('current_tenant_id') or getattr(g, 'current_tenant_id', None)
    if not tenant_id:
        flash('No organization selected', 'error')
        return redirect(url_for('main.dashboard'))

    try:
        inventory_items = db.session.execute(
            db.select(Inventory).where(Inventory.tenant_id == tenant_id)
        ).scalars().all()

        # Get recent transactions
        recent_transactions = db.session.execute(
            db.select(InventoryTransaction)
            .where(InventoryTransaction.tenant_id == tenant_id)
            .order_by(InventoryTransaction.created_at.desc())
            .limit(20)
        ).scalars().all()

        # Identify low stock items
        low_stock = [item for item in inventory_items
                     if item.quantity_on_hand <= item.reorder_level]

        from app.models.part import Part
        g.current_tenant_id = tenant_id
        parts = Part.get_all_sorted()
        return render_template('administrator/inventory.html',
                             inventory_items=inventory_items,
                             recent_transactions=recent_transactions,
                             low_stock=low_stock,
                             total_items=len(inventory_items),
                             parts=parts)

    except Exception as e:
        logger.error(f"Failed to load inventory: {e}")
        flash('Failed to load inventory', 'error')
        return render_template('administrator/inventory.html',
                             inventory_items=[],
                             recent_transactions=[],
                             low_stock=[],
                             total_items=0,
                             parts=[])


@administrator_bp.route('/inventory/adjust', methods=['POST'])
@handle_database_errors
def inventory_adjust():
    """Adjust inventory stock level"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    from app.models.inventory import Inventory, InventoryTransaction
    from app.extensions import db

    tenant_id = session.get('current_tenant_id') or getattr(g, 'current_tenant_id', None)
    if not tenant_id:
        flash('No organization selected', 'error')
        return redirect(url_for('main.dashboard'))

    part_id = request.form.get('part_id', type=int)
    inventory_id = request.form.get('inventory_id', type=int)
    adjustment = request.form.get('quantity', type=int)
    transaction_type = sanitize_input(request.form.get('transaction_type', 'adjustment'))
    notes = sanitize_input(request.form.get('notes', ''))

    if adjustment is None:
        flash('Érvénytelen mennyiség', 'error')
        return redirect(url_for('administrator.inventory'))

    try:
        # If part_id given, find or create inventory item
        if part_id and not inventory_id:
            item = db.session.execute(
                db.select(Inventory).where(
                    Inventory.tenant_id == tenant_id,
                    Inventory.part_id == part_id
                )
            ).scalar_one_or_none()
            if not item:
                item = Inventory(
                    tenant_id=tenant_id,
                    part_id=part_id,
                    quantity_on_hand=0,
                    reorder_level=5,
                )
                db.session.add(item)
                db.session.flush()
            inventory_id = item.inventory_id
        elif inventory_id:
            item = db.session.execute(db.select(Inventory).where(Inventory.inventory_id == inventory_id)).scalar_one_or_none()
            if not item or item.tenant_id != tenant_id:
                flash('Készlet elem nem található', 'error')
                return redirect(url_for('administrator.inventory'))
        else:
            flash('Hiányzó Termék adat', 'error')
            return redirect(url_for('administrator.inventory'))

        # Update quantity
        item.quantity_on_hand += adjustment

        # Record transaction
        transaction = InventoryTransaction(
            tenant_id=tenant_id,
            inventory_id=inventory_id,
            transaction_type=transaction_type,
            quantity=adjustment,
            performed_by=session.get('user_id'),
            notes=notes,
        )
        db.session.add(transaction)
        db.session.commit()

        flash(f'Inventory adjusted by {adjustment:+d} units', 'success')

    except Exception as e:
        logger.error(f"Failed to adjust inventory: {e}")
        db.session.rollback()
        flash('Failed to adjust inventory', 'error')

    return redirect(url_for('administrator.inventory'))


# =============================================================================
# SUBSCRIPTION MANAGEMENT
# =============================================================================

@administrator_bp.route('/subscription')
@handle_database_errors
@log_function_call
def subscription_management():
    """Subscription and plan management page"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    from app.models.subscription import Subscription
    from app.models.tenant import Tenant
    from app.extensions import db

    tenant_id = session.get('current_tenant_id') or getattr(g, 'current_tenant_id', None)
    if not tenant_id:
        flash('No organization selected', 'error')
        return redirect(url_for('main.dashboard'))

    try:
        tenant = Tenant.find_by_id(tenant_id)
        subscription = db.session.execute(
            db.select(Subscription).where(Subscription.tenant_id == tenant_id)
        ).scalar_one_or_none()

        current_plan = {
            'name': subscription.plan_name if subscription else 'Ingyenes',
            'status': subscription.status if subscription else 'active',
            'price': subscription.price if subscription else 0,
        }
        usage = {}
        return render_template('administrator/subscription.html',
                             tenant=tenant,
                             subscription=subscription,
                             current_plan=current_plan,
                             usage=usage)

    except Exception as e:
        logger.error(f"Failed to load subscription info: {e}")
        flash('Failed to load subscription information', 'error')
        return render_template('administrator/subscription.html',
                             tenant=None,
                             subscription=None,
                             current_plan={'name': 'Ingyenes', 'status': 'active'},
                             usage={})


# =============================================================================
# DB TOOLS — adatbázis állapot ellenőrzés és munkák visszaállítása
# =============================================================================

@administrator_bp.route('/db-tools')
@handle_database_errors
def db_tools():
    """Database status and recovery tool"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    from app.extensions import db
    from app.models.job import Job
    from app.models.customer import Customer
    from app.models.service import Service
    from app.models.part import Part

    try:
        tenant_id = session.get('current_tenant_id') or 1
        jobs = db.session.execute(db.select(Job).where(Job.tenant_id == tenant_id)).scalars().all()
        customers = db.session.execute(db.select(Customer).where(Customer.tenant_id == tenant_id)).scalars().all()
        services = db.session.execute(db.select(Service).where(Service.tenant_id == tenant_id)).scalars().all()
        parts = db.session.execute(db.select(Part).where(Part.tenant_id == tenant_id)).scalars().all()

        # notes oszlop ellenőrzése
        notes_col_exists = False
        try:
            db.session.execute(db.text("SELECT notes FROM job LIMIT 1"))
            notes_col_exists = True
        except Exception:
            db.session.rollback()

        stats = {
            'jobs': len(jobs),
            'customers': len(customers),
            'services': len(services),
            'parts': len(parts),
            'notes_col': notes_col_exists,
        }
        return render_template('administrator/db_tools.html', stats=stats, jobs=jobs, customers=customers)

    except Exception as e:
        logger.error(f"DB tools error: {e}")
        flash(f'Hiba: {e}', 'error')
        return redirect(url_for('administrator.dashboard'))


@administrator_bp.route('/db-tools/migrate', methods=['POST'])
@handle_database_errors
def db_migrate():
    """Run safe migrations (ADD COLUMN only, never drops)"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    from app.extensions import db
    results = []

    migrations = [
        ("notes oszlop (job)", "ALTER TABLE job ADD COLUMN IF NOT EXISTS notes TEXT"),
    ]

    for name, sql in migrations:
        try:
            db.session.execute(db.text(sql))
            db.session.commit()
            results.append(f"✅ {name}")
        except Exception as e:
            db.session.rollback()
            results.append(f"⚠️ {name}: {e}")

    flash(' | '.join(results), 'info')
    return redirect(url_for('administrator.db_tools'))


@administrator_bp.route('/db-tools/recover-job', methods=['POST'])
@handle_database_errors
def recover_job():
    """Manually re-create a lost job"""
    redirect_response = require_admin_login()
    if redirect_response:
        return redirect_response

    from app.extensions import db
    from app.models.job import Job
    from datetime import datetime

    try:
        customer_id = request.form.get('customer_id', type=int)
        job_date_str = request.form.get('job_date', '')
        notes = request.form.get('notes', '').strip()
        tenant_id = session.get('current_tenant_id') or 1

        if not customer_id or not job_date_str:
            flash('Ügyfél és dátum kötelező!', 'error')
            return redirect(url_for('administrator.db_tools'))

        job_date = datetime.strptime(job_date_str, '%Y-%m-%d').date()
        job = Job(
            customer=customer_id,
            job_date=job_date,
            tenant_id=tenant_id,
            total_cost=0.0,
            completed=False,
            paid=False,
            notes=notes or None,
        )
        db.session.add(job)
        db.session.commit()
        flash(f'✅ Munka #{job.job_id} visszaállítva!', 'success')

    except Exception as e:
        db.session.rollback()
        logger.error(f"Job recovery error: {e}")
        flash(f'Hiba: {e}', 'error')

    return redirect(url_for('administrator.db_tools'))

